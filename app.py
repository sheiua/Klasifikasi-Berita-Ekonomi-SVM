import streamlit as st
from openpyxl import load_workbook
import xlwt
import io
import joblib

# Load model SVM (pastikan modelnya sudah dilatih & disimpan sebagai .pkl)
@st.cache_resource
def load_model():
    return joblib.load("model_svm_berita.pkl")

model = load_model()

st.title("📑 Filter Berita Ekonomi")

uploaded_file = st.file_uploader("Upload file Berita Ekonomi (.xlsx)", type="xlsx")

if uploaded_file:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header = [str(h) for h in rows[0]]
    data_rows = rows[1:]

    idx_kolom5 = header.index("5. APA yang terjadi pada fenomena ekonomi yang ditemukan ?")

    hasil = []
    for row in data_rows:
        teks = str(row[idx_kolom5]) if row[idx_kolom5] is not None else ""
        pred = model.predict([teks])[0]
        if pred == 1:
            hasil.append(row)

    if hasil:
        wb_out = xlwt.Workbook()
        ws_out = wb_out.add_sheet("Berita_Ekonomi_Saja")

        for j, kol in enumerate(header):
            ws_out.write(0, j, kol)

        for i, baris in enumerate(hasil, 1):
            for j, val in enumerate(baris):
                ws_out.write(i, j, val)

        buf = io.BytesIO()
        wb_out.save(buf)
        buf.seek(0)

        st.download_button(
            "📥 Unduh Hanya Berita Ekonomi",
            data=buf,
            file_name="berita_ekonomi_saja.xls",
            mime="application/vnd.ms-excel"
        )
    else:
        st.warning("Tidak ada berita ekonomi yang terdeteksi dalam file ini!")
